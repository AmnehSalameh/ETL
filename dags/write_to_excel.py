import openpyxl

WorkingDirectory = "/opt/airflow/dags/w3c"
StarSchema = WorkingDirectory + "/StarSchema/"

def writeToExcel(sheet_name, data):
    # open output excel to write date dimension data into
    OutputExcelFile = StarSchema+ "FinalOutput.xlsx"  # Excel output file
    # Try to load the existing workbook; if not found, create a new one
    try:
        wb = openpyxl.load_workbook(OutputExcelFile)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
  
    # Remove the default "Sheet" if it exists and is empty
    if "Sheet" in wb.sheetnames and len(wb["Sheet"]["A"]) == 1:  # Only header row exists
        del wb["Sheet"]
        
    
    # If the sheet already exists, select it; otherwise, create a new one
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Append data to the selected sheet
    for row in data:
        ws.append(row)

    # Save the workbook
    wb.save(OutputExcelFile)
    wb.close()
